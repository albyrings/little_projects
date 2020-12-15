import openpyxl
import seaborn as sns
import numpy as np
_author_='Alberto Anelli'
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import matplotlib.patches as mpatches
def excel_array(vettore, foglio_lavoro, riga_inizio, riga_fine, colonna):
    for i in range(riga_inizio, riga_fine + 1):
               vettore.append(foglio_lavoro.cell(i, colonna).value)
    return np.array(vettore)

#input dati
wb1 = load_workbook('EEE.xlsx')
foglio = wb1['Foglio1']

x = []
y = []
z = []
V = excel_array(x, foglio,2,2017, 23)
I = excel_array(y, foglio,2,2017, 5)
Z = excel_array(z, foglio,2,2017, 3)
print(V)
print(x)
#grafico
plt.title('grafico rate')
plt.xlabel('data e ora')
plt.ylabel('rate')
red_patch = mpatches.Patch(color='red', label='rate totale')
green_patch = mpatches.Patch(color='blue', label='rate vero')
#ax=plt.subplot()

#genero un sistema di assi che condivide l'asse x con ax
#a2=ax.twinx()
#ax.plot(x,z)
plt.legend(handles=[red_patch,green_patch])
plt.plot(x,z,linestyle='solid',color='red',marker='')
plt.plot(x,y,linestyle='solid',color='blue',marker='')


plt.show()
